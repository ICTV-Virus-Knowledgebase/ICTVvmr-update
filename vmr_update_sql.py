#!/usr/bin/env python3
"""Generate SQL update and insert scripts from an ICTV VMR workbook."""

from __future__ import annotations

import argparse
import math
import re
import shlex
import subprocess
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from numbers import Integral, Real
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Set, Tuple

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_WORKBOOK = (
    Path(__file__).resolve().parent
    / "VMR_MSL40.v2.20251013.editor_dbs_20260202_v2.xlsx"
)
ERROR_FILENAME = "errors.xlsx"
DELETES_FILENAME = "vmr_1_deletes.sql"
UPDATES_FILENAME = "vmr_2_updates.sql"
INSERTS_FILENAME = "vmr_3_inserts.sql"
COLUMN_VALUE_INSERTS_FILENAME = "vmr_0_cv_inserts.sql"
UPDATE_SORTS_FILENAME = "vmr_4_update_sorts.sql"
QC_SPS_FILENAME = "vmr_5_exec_qc_sps.sql"
VERSION_FILE = Path("version_git.txt")

# Columns A:AG that must appear (in order) on the worksheets we process.
REQUIRED_COLUMNS: List[str] = [
    "Isolate ID",
    "Species Sort",
    "Isolate Sort",
    "Realm",
    "Subrealm",
    "Kingdom",
    "Subkingdom",
    "Phylum",
    "Subphylum",
    "Class",
    "Subclass",
    "Order",
    "Suborder",
    "Family",
    "Subfamily",
    "Genus",
    "Subgenus",
    "Species",
    "ICTV_ID",
    "Exemplar or additional isolate",
    "Virus name(s)",
    "Virus name abbreviation(s)",
    "Virus isolate designation",
    "Virus GENBANK accession",
    "Genome coverage",
    "Genome",
    "Host source",
    "Accessions Link",
    "Editor Notes",
    "QC_status",
    "QC_taxon_inher_molecule",
    "QC_taxon_change",
    "QC_taxon_proposal",
]

READ_ONLY_COLUMNS = {
    "Isolate ID",
    "Species Sort",
    "Realm",
    "Subrealm",
    "Kingdom",
    "Subkingdom",
    "Phylum",
    "Subphylum",
    "Class",
    "Subclass",
    "Order",
    "Suborder",
    "Family",
    "Subfamily",
    "Genus",
    "Subgenus",
    "ICTV_ID",
    "Accessions Link",
}

# Mapping of VMR columns to SQL columns when generating UPDATE statements.
UPDATABLE_TO_SQL = {
    "Isolate Sort": "isolate_sort",
    "Species": "species_name",
    "Exemplar or additional isolate": "isolate_type",
    "Virus name(s)": "isolate_names",
    "Virus name abbreviation(s)": "isolate_abbrevs",
    "Virus isolate designation": "isolate_designation",
    "Virus GENBANK accession": "genbank_accessions",
    "Genome coverage": "genome_coverage",
    "Genome": "molecule",
    "Host source": "host_source",
    "Editor Notes": "notes",
}

# Mapping used when creating INSERT statements for new records.
INSERT_COLUMN_MAPPING: Sequence[Tuple[str, str]] = (
    #    ("taxnode_id", "ICTV_ID"), # compute this from species_name, post-insert
    ("species_sort", "Species Sort"),
    ("isolate_sort", "Isolate Sort"),
    ("species_name", "Species"),
    ("isolate_type", "Exemplar or additional isolate"),
    ("isolate_names", "Virus name(s)"),
    ("isolate_abbrevs", "Virus name abbreviation(s)"),
    ("isolate_designation", "Virus isolate designation"),
    ("genbank_accessions", "Virus GENBANK accession"),
    ("genome_coverage", "Genome coverage"),
    ("molecule", "Genome"),
    ("host_source", "Host source"),
    ("notes", "Editor Notes"),
)

INT_COLUMNS = {"taxnode_id", "species_sort", "isolate_sort"}
INVALID_VALUE = object()
BLANK_CHECK_COLUMNS = REQUIRED_COLUMNS[:28]

ERROR_CONTEXT_COLUMNS: Sequence[Tuple[str, str]] = (
    ("Species Name", "Species"),
    ("ICTV_ID", "ICTV_ID"),
    ("Exemplar or additional isolate", "Exemplar or additional isolate"),
    ("Virus name(s)", "Virus name(s)"),
    ("Virus name abbreviation(s)", "Virus name abbreviation(s)"),
    ("Virus isolate designation", "Virus isolate designation"),
    ("Virus GENBANK accession", "Virus GENBANK accession"),
)


@dataclass
class ColumnConstraint:
    allowed_values: Set[str]
    canonical_map: Dict[str, str]


class ProcessingHalted(Exception):
    """Raised when validation should stop immediately."""


@dataclass
class ColumnValueInsertEntry:
    column: str
    value: str
    rows: List[int]


@dataclass(frozen=True)
class ColumnValueTarget:
    table: str
    columns: Tuple[str, ...]


COLUMN_VALUE_TARGETS = {
    "host source": ColumnValueTarget("taxonomy_host_source", ("host_source",)),
    "genome coverage": ColumnValueTarget(
        "taxonomy_genome_coverage", ("genome_coverage",)
    ),
    "genome": ColumnValueTarget("taxonomy_molecule", ("abbrev", "name")),
}


@dataclass(frozen=True)
class TaxonomyReferencePaths:
    genome_coverage: Optional[Path]
    molecule: Optional[Path]
    host_source: Optional[Path]


@dataclass
class ErrorEntry:
    filename: str
    worksheet: str
    row: Optional[int]
    message: str
    severity: str


@dataclass
class UpdateEntry:
    isolate_id: str
    numeric_id: int
    row_number: int
    assignments: List[Tuple[str, Optional[object]]]
    original_values: List[Tuple[str, Optional[object]]]


@dataclass
class InsertEntry:
    row_number: int
    values: List[Tuple[str, Optional[object]]]


@dataclass
class DeleteEntry:
    isolate_id: str
    numeric_id: int
    target_value: str
    row_number: int
    details: List[Tuple[str, Optional[object]]]


@dataclass
class ProcessResult:
    updated_sheet: Optional[str]
    delete_entries: List[DeleteEntry]
    update_entries: List[UpdateEntry]
    insert_entries: List[InsertEntry]
    column_value_inserts: List[ColumnValueInsertEntry]


class ErrorCollector:
    """Collects errors and enforces the stop/continue policy."""

    def __init__(self, keep_going: bool, command: str, version: str, run_date: str) -> None:
        self.keep_going = keep_going
        self.entries: List[ErrorEntry] = []
        self.command = command
        self.version = version
        self.run_date = run_date
        self.row_context: Dict[Tuple[str, int], Dict[str, object]] = {}

    def add(
        self,
        filename: str,
        worksheet: str,
        row: Optional[int],
        message: str,
        *,
        severity: str = "ERROR",
    ) -> None:
        entry = ErrorEntry(filename, worksheet, row, message, severity)
        self.entries.append(entry)
        location = filename
        if worksheet:
            location += f"::{worksheet}"
        if row is not None:
            location += f" row {row}"
        print(f"{severity.upper()}: {location} - {message}", file=sys.stderr)
        if severity.upper() == "ERROR" and not self.keep_going:
            raise ProcessingHalted(message)

    def has_errors(self) -> bool:
        return any(entry.severity.upper() == "ERROR" for entry in self.entries)

    def extend_with_exception(self, filename: str, exc: Exception) -> None:
        self.entries.append(
            ErrorEntry(filename, "", None, f"Unhandled exception: {exc!r}", "ERROR")
        )
        print(f"ERROR: {filename} - Unhandled exception: {exc!r}", file=sys.stderr)

    def register_row_context(
        self,
        worksheet: str,
        row_number: int,
        values: Dict[str, object],
        changes: Dict[str, bool],
        isolate_id: Optional[str],
    ) -> None:
        self.row_context[(worksheet, row_number)] = {
            "values": values,
            "changes": changes,
            "isolate_id": isolate_id,
        }

    def write_excel(self, output_path: Path) -> None:
        context_headers = [display for display, _ in ERROR_CONTEXT_COLUMNS]
        rows: List[Dict[str, object]] = []
        change_flags: List[Dict[str, bool]] = []
        for entry in self.entries:
            row_data: Dict[str, object] = {
                "filename": entry.filename,
                "worksheet": entry.worksheet,
                "row": entry.row,
                "message": entry.message,
                "severity": entry.severity,
                "command": self.command,
                "version": self.version,
                "run_date": self.run_date,
            }
            context = None
            if entry.row is not None:
                context = self.row_context.get((entry.worksheet, int(entry.row)))
            isolate_id_value = None
            if context:
                isolate_id_value = context.get("isolate_id")
            row_data["Isolate ID"] = isolate_id_value
            context_changes: Dict[str, bool] = {}
            for display in context_headers:
                value = None
                changed = False
                if context:
                    value = context["values"].get(display)
                    changed = bool(context["changes"].get(display, False))
                row_data[display] = value
                context_changes[display] = changed
            rows.append(row_data)
            change_flags.append(context_changes)

        columns = [
            "filename",
            "worksheet",
            "row",
            "Isolate ID",
            *context_headers,
            "message",
            "severity",
            "command",
            "version",
            "run_date",
        ]
        df = pd.DataFrame(rows, columns=columns)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            sheet_name = "Sheet1"
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            start_row = 2
            green_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
            ictv_col_idx = columns.index("ICTV_ID") + 1 if "ICTV_ID" in columns else None
            accession_col_idx = (
                columns.index("Virus GENBANK accession") + 1
                if "Virus GENBANK accession" in columns
                else None
            )
            isolate_col_idx = columns.index("Isolate ID") + 1 if "Isolate ID" in columns else None
            for row_offset, flags in enumerate(change_flags):
                excel_row = start_row + row_offset
                for display in context_headers:
                    excel_col = columns.index(display) + 1
                    if flags.get(display):
                        worksheet.cell(row=excel_row, column=excel_col).fill = green_fill
                if ictv_col_idx is not None:
                    cell = worksheet.cell(row=excel_row, column=ictv_col_idx)
                    value = cell.value
                    if value is not None:
                        stripped = str(value).strip()
                        if stripped:
                            cell.hyperlink = f"https://ictv.global/id/{stripped}"
                            cell.style = "Hyperlink"
                if isolate_col_idx is not None:
                    cell = worksheet.cell(row=excel_row, column=isolate_col_idx)
                    value = cell.value
                    if value is not None:
                        stripped = str(value).strip()
                        if stripped:
                            upper_value = stripped.upper()
                            numeric_part = (
                                upper_value[3:] if upper_value.startswith("VMR") else upper_value
                            )
                            hyperlink = f"https://ictv.global/id/VMR{numeric_part}"
                            cell.hyperlink = hyperlink
                            cell.style = "Hyperlink"
                if accession_col_idx is not None:
                    cell = worksheet.cell(row=excel_row, column=accession_col_idx)
                    value = cell.value
                    if isinstance(value, str):
                        entries = split_accession_entries(value)
                        tokens = []
                        for _, accession in entries:
                            cleaned = accession.replace(" ", "")
                            if cleaned:
                                tokens.append(cleaned)
                        if tokens:
                            joined = ",".join(tokens)
                            cell.hyperlink = (
                                f"https://www.ncbi.nlm.nih.gov/nuccore/{joined}"
                            )
                            cell.style = "Hyperlink"

            for col_idx, column_name in enumerate(columns, start=1):
                max_length = len(str(column_name))
                for cell in worksheet.iter_rows(
                    min_row=1,
                    max_row=worksheet.max_row,
                    min_col=col_idx,
                    max_col=col_idx,
                ):
                    value = cell[0].value
                    if value is None:
                        continue
                    max_length = max(max_length, len(str(value)))
                worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Validate a VMR workbook and generate MariaDB SQL update and insert scripts."
        )
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        default=str(DEFAULT_WORKBOOK),
        help="Path to the VMR Excel workbook (default: %(default)s)",
    )
    parser.add_argument(
        "-o",
        "--output-dir",
        help=(
            "Directory where SQL and error reports are written. Default is the"
            " workbook filename without the .xlsx suffix."
        ),
    )
    parser.add_argument(
        "-k",
        "--keep-going",
        action="store_true",
        help="Continue processing after encountering validation errors.",
    )
    parser.add_argument(
        "--strict-accession",
        action="store_true",
        help=(
            "Emit warnings for accession changes that only adjust whitespace, "
            "segment labels, or populate previously empty values."
        ),
    )
    parser.add_argument(
        "--updates-sql",
        default=UPDATES_FILENAME,
        help="Filename for UPDATE statements (default: %(default)s)",
    )
    parser.add_argument(
        "--deletes-sql",
        default=DELETES_FILENAME,
        help="Filename for DELETE statements (default: %(default)s)",
    )
    parser.add_argument(
        "--inserts-sql",
        default=INSERTS_FILENAME,
        help="Filename for INSERT statements (default: %(default)s)",
    )
    parser.add_argument(
        "--column-values-sql",
        default=COLUMN_VALUE_INSERTS_FILENAME,
        help="Filename for column value INSERT statements (default: %(default)s)",
    )
    parser.add_argument(
        "--update-sorts-sql",
        default=UPDATE_SORTS_FILENAME,
        help="Filename for the update-sorts stored procedure call (default: %(default)s)",
    )
    parser.add_argument(
        "--qc-sps-sql",
        default=QC_SPS_FILENAME,
        help="Filename for the QC stored procedure call (default: %(default)s)",
    )
    parser.add_argument(
        "--errors-xlsx",
        default=ERROR_FILENAME,
        help="Filename for the error report workbook (default: %(default)s)",
    )
    parser.add_argument(
        "--vmr-export",
        help=(
            "Optional path to a vmr_export view export (CSV/XLSX) to validate "
            "against the workbook 'Original' worksheet."
        ),
    )
    parser.add_argument(
        "--taxonomy-genome-coverage",
        help=(
            "Optional path to a taxonomy_genome_coverage table export (CSV/XLSX) "
            "for validating 'Original Column Values'."
        ),
    )
    parser.add_argument(
        "--taxonomy-molecule",
        help=(
            "Optional path to a taxonomy_molecule table export (CSV/XLSX) "
            "for validating 'Original Column Values'."
        ),
    )
    parser.add_argument(
        "--taxonomy-host-source",
        help=(
            "Optional path to a taxonomy_host_source table export (CSV/XLSX) "
            "for validating 'Original Column Values'."
        ),
    )
    return parser.parse_args()


def read_version() -> str:
    if VERSION_FILE.exists():
        return VERSION_FILE.read_text(encoding="utf-8").strip()
    fallback = Path("version.txt")
    if fallback.exists():
        return fallback.read_text(encoding="utf-8").strip()
    return "unknown"


def normalize_string(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    if isinstance(value, bytes):
        stripped = value.decode("utf-8", errors="ignore").strip()
        return stripped or None
    if isinstance(value, Real):
        if isinstance(value, float) and math.isnan(value):
            return None
        return str(value)
    return str(value).strip() or None


def normalize_column_key(value: object) -> Optional[str]:
    text = normalize_string(value)
    if text is None:
        return None
    return text.lower()


def parse_hyperlink_label(value: object) -> Optional[str]:
    text = normalize_string(value)
    if text is None:
        return None
    match = re.fullmatch(r'=HYPERLINK\(".*?","(.*)"\)', text, flags=re.IGNORECASE)
    if match:
        return match.group(1)
    return None


def normalize_hyperlink_label(value: object) -> Optional[str]:
    label = parse_hyperlink_label(value)
    if label is not None:
        return normalize_string(label)
    return normalize_string(value)


def read_tabular_file(path: Path, *, sheet_name: Optional[str] = None) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        sep = "\t" if suffix == ".txt" else ","
        return pd.read_csv(path, sep=sep)
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(path, sheet_name=sheet_name, header=0, engine="openpyxl")
    raise ValueError(f"Unsupported file format for {path}")


def select_excel_sheet(path: Path, preferred: Sequence[str]) -> Optional[str]:
    try:
        with pd.ExcelFile(path) as xl:
            for name in preferred:
                if name in xl.sheet_names:
                    return name
            return xl.sheet_names[0] if xl.sheet_names else None
    except Exception:
        return None


def canonicalize_column_value(value: str) -> str:
    return "".join(ch for ch in value if ch.isalnum()).lower()


def is_abolish_value(value: object) -> bool:
    text = normalize_string(value)
    return text is not None and text.lower() == "abolish"


def normalize_int_like(value: object) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, Integral):
        return int(value)
    if isinstance(value, Real):
        if math.isnan(value):
            return None
        if float(value).is_integer():
            return int(value)
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            number = float(stripped)
        except ValueError:
            return None
        if math.isnan(number):
            return None
        if number.is_integer():
            return int(number)
        return None
    return None


def normalize_isolate_type(value: object) -> Optional[str]:
    text = normalize_string(value)
    if text is None:
        return None
    text = text.upper()
    if text in {"E", "A"}:
        return text
    if text.startswith("EXEM"):
        return "E"
    if text.startswith("ADD"):
        return "A"
    return None


def build_column_lookup(df: pd.DataFrame) -> Dict[str, str]:
    lookup: Dict[str, str] = {}
    for name in df.columns:
        key = normalize_column_key(name)
        if key and key not in lookup:
            lookup[key] = name
    return lookup


def load_reference_values(
    path: Path,
    *,
    preferred_sheets: Sequence[str],
    column_candidates: Sequence[str],
    errors: ErrorCollector,
) -> Optional[Set[str]]:
    if not path.exists():
        errors.add(path.name, "", None, f"Reference file not found: {path}")
        return None
    sheet = None
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        sheet = select_excel_sheet(path, preferred_sheets)
        if sheet is None:
            errors.add(path.name, "", None, "Reference workbook does not contain any sheets.")
            return None
    try:
        df = read_tabular_file(path, sheet_name=sheet)
    except Exception as exc:
        errors.add(path.name, "", None, f"Failed to read reference file: {exc!r}")
        return None
    lookup = build_column_lookup(df)
    target_column = None
    for candidate in column_candidates:
        col = lookup.get(candidate.lower())
        if col is not None:
            target_column = col
            break
    if target_column is None:
        errors.add(
            path.name,
            "",
            None,
            "Reference file is missing required column(s): " + ", ".join(column_candidates),
        )
        return None
    values: Set[str] = set()
    for cell in df[target_column]:
        text = normalize_string(cell)
        if text is not None:
            values.add(text)
    return values


def load_taxonomy_reference_sets(
    paths: TaxonomyReferencePaths,
    errors: ErrorCollector,
) -> Dict[str, Set[str]]:
    reference_sets: Dict[str, Set[str]] = {}
    if paths.genome_coverage:
        values = load_reference_values(
            paths.genome_coverage,
            preferred_sheets=["taxonomy_genome_coverage", "genome_coverage"],
            column_candidates=["name", "genome_coverage"],
            errors=errors,
        )
        if values:
            reference_sets["Genome coverage"] = values
    if paths.molecule:
        values = load_reference_values(
            paths.molecule,
            preferred_sheets=["taxonomy_molecule", "molecule"],
            column_candidates=["abbrev", "name"],
            errors=errors,
        )
        if values:
            reference_sets["Genome"] = values
    if paths.host_source:
        values = load_reference_values(
            paths.host_source,
            preferred_sheets=["taxonomy_host_source", "host_source"],
            column_candidates=["host_source"],
            errors=errors,
        )
        if values:
            reference_sets["Host source"] = values
    return reference_sets


def check_original_column_values_against_taxonomy(
    original_values_df: pd.DataFrame,
    workbook_name: str,
    errors: ErrorCollector,
    reference_sets: Dict[str, Set[str]],
) -> None:
    if not reference_sets:
        return
    for column, allowed_values in reference_sets.items():
        if column not in original_values_df.columns:
            errors.add(
                workbook_name,
                "Original Column Values",
                None,
                f"Column '{column}' missing from 'Original Column Values'.",
            )
            continue
        for idx, cell in original_values_df[column].items():
            text = normalize_string(cell)
            if text is None:
                continue
            if text not in allowed_values:
                errors.add(
                    workbook_name,
                    "Original Column Values",
                    excel_row(idx),
                    (
                        f"Value '{text}' in column '{column}' is not present in the "
                        "corresponding taxonomy reference table."
                    ),
                )


def normalize_vmr_export_value(column: str, value: object) -> Optional[str]:
    if column == "Virus isolate designation":
        if isinstance(value, (pd.Timestamp, datetime)):
            return value.strftime("%b-%y")
    if column in {"Isolate ID", "ICTV_ID", "Accessions Link", "QC_taxon_proposal"}:
        return normalize_hyperlink_label(value)
    if column == "Exemplar or additional isolate":
        normalized = normalize_isolate_type(value)
        return normalized if normalized is not None else normalize_string(value)
    if column in {"Isolate Sort", "Species Sort"}:
        number = normalize_int_like(value)
        return str(number) if number is not None else normalize_string(value)
    return normalize_string(value)


def compare_original_to_vmr_export(
    original_df: pd.DataFrame,
    workbook_name: str,
    errors: ErrorCollector,
    vmr_export_path: Path,
) -> None:
    if not vmr_export_path.exists():
        errors.add(vmr_export_path.name, "", None, f"VMR export file not found: {vmr_export_path}")
        return
    sheet = None
    if vmr_export_path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        sheet = select_excel_sheet(
            vmr_export_path, ["vmr_export", "vmr export", "VMR export"]
        )
    try:
        export_df = read_tabular_file(vmr_export_path, sheet_name=sheet)
    except Exception as exc:
        errors.add(vmr_export_path.name, "", None, f"Failed to read VMR export: {exc!r}")
        return

    export_lookup = build_column_lookup(export_df)
    missing_columns = [
        name for name in REQUIRED_COLUMNS if name.lower() not in export_lookup
    ]
    if missing_columns:
        errors.add(
            vmr_export_path.name,
            "",
            None,
            "VMR export is missing required columns: " + ", ".join(missing_columns),
        )
        return

    export_rows: Dict[str, Tuple[int, pd.Series]] = {}
    for idx, row in export_df.iterrows():
        raw_isolate = row[export_lookup["isolate id"]]
        isolate_label = normalize_hyperlink_label(raw_isolate)
        isolate_id = normalize_isolate_id(isolate_label)
        if isolate_id is None:
            continue
        if isolate_id in export_rows:
            errors.add(
                vmr_export_path.name,
                "vmr_export",
                excel_row(idx),
                f"Isolate ID {isolate_id} appears multiple times in vmr_export.",
            )
            continue
        export_rows[isolate_id] = (excel_row(idx), row)

    original_map = original_df[original_df["__isolate_id"].notna()].set_index("__isolate_id")
    original_ids = set(original_map.index)
    export_ids = set(export_rows.keys())

    for isolate_id in sorted(export_ids - original_ids):
        export_row_num = export_rows[isolate_id][0]
        errors.add(
            vmr_export_path.name,
            "vmr_export",
            export_row_num,
            f"Isolate ID {isolate_id} missing from Original worksheet.",
        )

    for isolate_id in sorted(original_ids - export_ids):
        row_number = int(original_map.loc[isolate_id]["__row_number"])
        errors.add(
            workbook_name,
            "Original",
            row_number,
            f"Isolate ID {isolate_id} missing from vmr_export.",
        )

    for isolate_id in sorted(original_ids & export_ids):
        orig_row = original_map.loc[isolate_id]
        if isinstance(orig_row, pd.DataFrame):
            orig_row = orig_row.iloc[0]
        export_row_num, export_row = export_rows[isolate_id]
        for column in REQUIRED_COLUMNS:
            export_value = export_row[export_lookup[column.lower()]]
            orig_norm = normalize_vmr_export_value(column, orig_row[column])
            export_norm = normalize_vmr_export_value(column, export_value)
            if orig_norm != export_norm:
                errors.add(
                    workbook_name,
                    "Original",
                    int(orig_row["__row_number"]),
                    (
                        f"Column '{column}' for isolate {isolate_id} does not match "
                        f"vmr_export (Original='{orig_norm or ''}', vmr_export='{export_norm or ''}')."
                    ),
                )


def values_equal(original: object, updated: object, column: str) -> bool:
    if column == "Exemplar or additional isolate":
        return normalize_isolate_type(original) == normalize_isolate_type(updated)
    if column in {"Isolate Sort", "Species Sort"}:
        return normalize_int_like(original) == normalize_int_like(updated)
    return normalize_string(original) == normalize_string(updated)


def normalize_isolate_id(value: object) -> Optional[str]:
    text = normalize_string(value)
    if not text:
        return None
    return text.upper()


def extract_isolate_numeric(isolate_id: str) -> Optional[int]:
    match = re.fullmatch(r"VMR(\d+)", isolate_id)
    if not match:
        return None
    return int(match.group(1))


def excel_row(index: int) -> int:
    return int(index) + 2


def validate_headers(
    filename: str,
    sheet_name: str,
    actual_columns: Sequence[str],
    errors: ErrorCollector,
) -> None:
    expected_len = len(REQUIRED_COLUMNS)
    if len(actual_columns) < expected_len:
        missing = REQUIRED_COLUMNS[len(actual_columns) :]
        errors.add(
            filename,
            sheet_name,
            None,
            "Worksheet is missing required columns: " + ", ".join(missing),
        )
        return
    for idx, expected in enumerate(REQUIRED_COLUMNS):
        actual = actual_columns[idx]
        if actual != expected:
            errors.add(
                filename,
                sheet_name,
                None,
                f"Column {idx + 1} mismatch: expected '{expected}' but found '{actual}'",
            )


def prepare_dataframe(workbook: Path, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(workbook, sheet_name=sheet_name, usecols=REQUIRED_COLUMNS)
    df = df.copy()
    blank_mask = (
        df[BLANK_CHECK_COLUMNS]
        .map(lambda value: normalize_string(value) is None)
        .all(axis=1)
    )
    if blank_mask.any():
        df = df.loc[~blank_mask].copy()
    df["__row_number"] = (df.index + 2).astype(int)
    df["__isolate_id"] = df["Isolate ID"].apply(normalize_isolate_id)
    return df


def build_column_value_maps(df: pd.DataFrame) -> Dict[str, Dict[str, List[int]]]:
    column_map: Dict[str, Dict[str, List[int]]] = {}
    for raw_column in df.columns:
        if pd.isna(raw_column):
            continue
        column_name = normalize_string(raw_column)
        if column_name is None:
            continue
        values: Dict[str, List[int]] = {}
        for idx, cell in df[raw_column].items():
            text = normalize_string(cell)
            if text is None:
                continue
            values.setdefault(text, []).append(excel_row(idx))
        if values:
            column_map[column_name] = values
    return column_map


def check_column_value_revisions(
    workbook: Path,
    workbook_name: str,
    updated_df: pd.DataFrame,
    errors: ErrorCollector,
) -> Dict[str, Dict[str, List[int]]]:
    new_values: Dict[str, Dict[str, List[int]]] = {}
    try:
        original_df = pd.read_excel(
            workbook, sheet_name="Original Column Values", header=0
        )
    except ValueError:
        errors.add(
            workbook_name,
            "Original Column Values",
            None,
            "Workbook does not contain an 'Original Column Values' worksheet.",
        )
        return {}

    updated_map = build_column_value_maps(updated_df)
    original_map = build_column_value_maps(original_df)

    for column, values in updated_map.items():
        original_values = original_map.get(column, {})
        if not original_values:
            for value, rows in values.items():
                errors.add(
                    workbook_name,
                    "Column Values",
                    rows[0],
                    (
                        f"Value '{value}' in column '{column}' does not appear on 'Original "
                        "Column Values'."
                    ),
                )
                new_values.setdefault(column, {})[value] = rows
            continue
        for value, rows in values.items():
            if value not in original_values:
                errors.add(
                    workbook_name,
                    "Column Values",
                    rows[0],
                    (
                        f"Value '{value}' in column '{column}' does not appear on 'Original "
                        "Column Values'."
                    ),
                )
                new_values.setdefault(column, {})[value] = rows

    for column, values in original_map.items():
        updated_values = updated_map.get(column, {})
        if not updated_values:
            for value, rows in values.items():
                errors.add(
                    workbook_name,
                    "Original Column Values",
                    rows[0],
                    (
                        f"Value '{value}' in column '{column}' is missing from 'Column Values'."
                    ),
                )
            continue
        for value, rows in values.items():
            if value not in updated_values:
                errors.add(
                    workbook_name,
                    "Original Column Values",
                    rows[0],
                    (
                        f"Value '{value}' in column '{column}' is missing from 'Column Values'."
                    ),
                )

    return new_values


def read_column_value_constraints(
    workbook: Path, workbook_name: str, errors: ErrorCollector
) -> Tuple[Dict[str, ColumnConstraint], List[ColumnValueInsertEntry]]:
    try:
        column_values_df = pd.read_excel(workbook, sheet_name="Column Values", header=0)
    except ValueError:
        errors.add(
            workbook_name,
            "Column Values",
            None,
            "Workbook does not contain a 'Column Values' worksheet; column value validation skipped.",
            severity="WARNING",
        )
        return {}, []

    new_column_values = check_column_value_revisions(
        workbook, workbook_name, column_values_df, errors
    )

    constraints: Dict[str, ColumnConstraint] = {}
    for raw_column in column_values_df.columns:
        if pd.isna(raw_column):
            continue
        column_name = normalize_string(raw_column)
        if column_name is None:
            continue
        allowed_values: Set[str] = set()
        canonical_map: Dict[str, str] = {}
        for cell in column_values_df[raw_column]:
            if pd.isna(cell):
                continue
            value = normalize_string(cell)
            if value is None:
                continue
            allowed_values.add(value)
            canonical = canonicalize_column_value(value)
            if canonical not in canonical_map:
                canonical_map[canonical] = value
        if allowed_values:
            constraints[column_name] = ColumnConstraint(allowed_values, canonical_map)
    column_value_inserts: List[ColumnValueInsertEntry] = []
    for column in sorted(new_column_values.keys(), key=lambda name: name.lower()):
        values = new_column_values[column]
        for value, rows in sorted(
            values.items(), key=lambda item: item[0].lower() if isinstance(item[0], str) else str(item[0])
        ):
            column_value_inserts.append(
                ColumnValueInsertEntry(column=column, value=value, rows=sorted(set(rows)))
            )
    return constraints, column_value_inserts


def check_column_value_constraints(
    updated_df: pd.DataFrame,
    workbook_name: str,
    updated_sheet: str,
    errors: ErrorCollector,
    constraints: Dict[str, ColumnConstraint],
) -> None:
    if not constraints:
        return
    for column, constraint in constraints.items():
        if column not in updated_df.columns:
            continue
        for idx, row in updated_df.iterrows():
            row_number = int(row["__row_number"])
            cell_value = row[column]
            if pd.isna(cell_value):
                value_text = None
            else:
                value_text = normalize_string(cell_value)
            if value_text is None:
                errors.add(
                    workbook_name,
                    updated_sheet,
                    row_number,
                    f"Column '{column}' must not be blank; select a value from 'Column Values'.",
                    severity="WARNING",
                )
                continue
            canonical = canonicalize_column_value(value_text)
            corrected = constraint.canonical_map.get(canonical)
            matches_allowed = value_text in constraint.allowed_values
            if matches_allowed and corrected is not None:
                if not (
                    isinstance(row[column], str)
                    and row[column] != corrected
                    and corrected == value_text
                ):
                    continue
            if corrected is not None:
                if corrected != value_text or (
                    isinstance(row[column], str) and row[column] != corrected
                ):
                    errors.add(
                        workbook_name,
                        updated_sheet,
                        row_number,
                        (
                            f"Column '{column}' value '{value_text}' adjusted to "
                            f"'{corrected}' to match 'Column Values'."
                        ),
                        severity="WARNING",
                    )
                updated_df.at[idx, column] = corrected
                continue
            errors.add(
                workbook_name,
                updated_sheet,
                row_number,
                f"Column '{column}' contains '{value_text}' which is not listed in 'Column Values'.",
            )


def determine_updated_sheet(
    sheet_names: Sequence[str], workbook_name: str, errors: ErrorCollector
) -> Optional[str]:
    pattern = re.compile(r"VMR MSL\d+")
    matches = [name for name in sheet_names if pattern.fullmatch(name)]
    if not matches:
        errors.add(
            workbook_name,
            "",
            None,
            "Workbook does not contain a worksheet named like 'VMR MSL[0-9]+'.",
        )
        return None
    if len(matches) > 1:
        errors.add(
            workbook_name,
            "",
            None,
            "Multiple updated worksheets found: " + ", ".join(matches),
        )
    return matches[0]


def check_original_ids(original_df: pd.DataFrame, workbook_name: str, errors: ErrorCollector) -> None:
    for _, row in original_df.iterrows():
        if row["__isolate_id"] is None:
            errors.add(
                workbook_name,
                "Original",
                int(row["__row_number"]),
                "Original worksheet contains a blank Isolate ID.",
            )


def check_isolate_ids(
    updated_df: pd.DataFrame,
    original_df: pd.DataFrame,
    workbook_name: str,
    updated_sheet: str,
    errors: ErrorCollector,
) -> None:
    updated_existing = updated_df[updated_df["__isolate_id"].notna()]
    original_existing = original_df[original_df["__isolate_id"].notna()]

    for df, sheet in ((updated_existing, updated_sheet), (original_existing, "Original")):
        for _, row in df.iterrows():
            isolate_id = row["__isolate_id"]
            if isolate_id and not re.fullmatch(r"VMR\d+", isolate_id):
                errors.add(
                    workbook_name,
                    sheet,
                    int(row["__row_number"]),
                    f"Invalid Isolate ID format: {row['Isolate ID']}",
                )

    for df, sheet in ((updated_existing, updated_sheet), (original_existing, "Original")):
        counts = df["__isolate_id"].value_counts()
        for isolate_id, count in counts.items():
            if count > 1:
                rows = df.loc[df["__isolate_id"] == isolate_id, "__row_number"].astype(int).tolist()
                rows_str = ", ".join(str(r) for r in rows)
                errors.add(
                    workbook_name,
                    sheet,
                    rows[0],
                    f"Isolate ID {isolate_id} appears multiple times (rows {rows_str}).",
                )

    updated_ids = set(updated_existing["__isolate_id"])
    original_ids = set(original_existing["__isolate_id"])

    for isolate_id in sorted(updated_ids - original_ids):
        row_num = int(
            updated_existing.loc[updated_existing["__isolate_id"] == isolate_id, "__row_number"].iloc[0]
        )
        errors.add(
            workbook_name,
            updated_sheet,
            row_num,
            f"Isolate ID {isolate_id} not present in Original worksheet.",
        )

    for isolate_id in sorted(original_ids - updated_ids):
        row_num = int(
            original_existing.loc[original_existing["__isolate_id"] == isolate_id, "__row_number"].iloc[0]
        )
        errors.add(
            workbook_name,
            "Original",
            row_num,
            f"Isolate ID {isolate_id} missing from updated worksheet.",
        )


def enforce_read_only(
    updated_df: pd.DataFrame,
    original_df: pd.DataFrame,
    workbook_name: str,
    updated_sheet: str,
    errors: ErrorCollector,
    abolished_ids: set[str],
) -> None:
    updated_map = updated_df.set_index("__isolate_id")
    original_map = original_df.set_index("__isolate_id")
    for isolate_id, orig_row in original_map.iterrows():
        if isolate_id not in updated_map.index or isolate_id is None:
            continue
        upd_row = updated_map.loc[isolate_id]
        if isinstance(upd_row, pd.DataFrame):
            continue
        if isolate_id in abolished_ids:
            continue
        for column in READ_ONLY_COLUMNS:
            if column == "Isolate ID":
                continue
            if not values_equal(orig_row[column], upd_row[column], column):
                errors.add(
                    workbook_name,
                    updated_sheet,
                    int(upd_row["__row_number"]),
                    f"Read-only column '{column}' changed for isolate {isolate_id}.",
                )


def split_accession_entries(value: object) -> List[Tuple[Optional[str], str]]:
    """Return segment/accession pairs extracted from a worksheet cell."""

    text = normalize_string(value)
    if not text:
        return []
    entries: List[Tuple[Optional[str], str]] = []
    for fragment in re.split(r"[;\n]+", text):
        part = fragment.strip()
        if not part:
            continue
        segment: Optional[str] = None
        accession = part
        if ":" in part:
            segment_text, accession_text = part.split(":", 1)
            segment = segment_text.strip() or None
            accession = accession_text.strip()
        if accession:
            entries.append((segment, accession))
    return entries


def canonicalize_accession_entries(
    entries: List[Tuple[Optional[str], str]], include_segment_names: bool
) -> List[Tuple[str, str]]:
    """Build a normalized representation of accession entries."""

    canonical: List[Tuple[str, str]] = []
    for segment, accession in entries:
        accession_norm = accession.upper()
        if include_segment_names:
            segment_norm = (segment or "").upper()
        else:
            segment_norm = ""
        canonical.append((segment_norm, accession_norm))
    return canonical


def classify_accession_change(original: object, updated: object) -> str:
    """Classify how the accession field changed between two values."""

    orig_entries = split_accession_entries(original)
    upd_entries = split_accession_entries(updated)

    if not orig_entries:
        if not upd_entries:
            return "whitespace"
        if normalize_string(original) is None:
            return "was_empty"
    if not upd_entries:
        if orig_entries:
            return "meaningful"
        return "whitespace"

    if canonicalize_accession_entries(orig_entries, True) == canonicalize_accession_entries(
        upd_entries, True
    ):
        return "whitespace"
    if canonicalize_accession_entries(orig_entries, False) == canonicalize_accession_entries(
        upd_entries, False
    ):
        return "segment_name"
    return "meaningful"


def parse_accession_tokens(value: object) -> List[str]:
    return [accession.upper() for _, accession in split_accession_entries(value)]


def check_new_record_accessions(
    updated_df: pd.DataFrame,
    workbook_name: str,
    updated_sheet: str,
    errors: ErrorCollector,
) -> None:
    new_rows = updated_df[
        (updated_df["__isolate_id"].isna()) & (~updated_df["__abolished"])
    ]
    if new_rows.empty:
        return

    existing_map: dict[str, int] = {}
    existing_rows = updated_df[
        (updated_df["__isolate_id"].notna()) & (~updated_df["__abolished"])
    ]
    for _, row in existing_rows.iterrows():
        row_number = int(row["__row_number"])
        for token in parse_accession_tokens(row["Virus GENBANK accession"]):
            existing_map.setdefault(token, row_number)

    seen_new: dict[str, int] = {}
    for _, row in new_rows.iterrows():
        row_number = int(row["__row_number"])
        tokens = parse_accession_tokens(row["Virus GENBANK accession"])
        duplicates = sorted({token for token in tokens if token in existing_map})
        if duplicates:
            refs = [
                f"{token} (worksheet row {existing_map[token]})" for token in duplicates
            ]
            errors.add(
                workbook_name,
                updated_sheet,
                row_number,
                "New record reuses existing accession(s): " + ", ".join(refs),
            )
            continue
        duplicates_new = sorted({token for token in tokens if token in seen_new})
        if duplicates_new:
            refs = [
                f"{token} (worksheet row {seen_new[token]})" for token in duplicates_new
            ]
            errors.add(
                workbook_name,
                updated_sheet,
                row_number,
                "New record reuses accession(s) from other new rows: "
                + ", ".join(refs),
            )
            continue
        for token in tokens:
            seen_new.setdefault(token, row_number)


def register_error_context(
    errors: ErrorCollector,
    updated_df: pd.DataFrame,
    original_df: pd.DataFrame,
    updated_sheet: str,
) -> None:
    if updated_sheet is None:
        return
    original_map = (
        original_df[original_df["__isolate_id"].notna()]
        .set_index("__isolate_id")
        if not original_df.empty
        else pd.DataFrame()
    )
    for _, row in updated_df.iterrows():
        row_number = int(row["__row_number"])
        isolate_id = row["__isolate_id"]
        orig_row: Optional[pd.Series]
        orig_row = None
        if isinstance(original_map, pd.DataFrame) and not original_map.empty and isolate_id:
            try:
                candidate = original_map.loc[isolate_id]
            except KeyError:
                candidate = None
            if candidate is not None:
                if isinstance(candidate, pd.DataFrame):
                    if not candidate.empty:
                        orig_row = candidate.iloc[0]
                else:
                    orig_row = candidate
        context_values: Dict[str, object] = {}
        context_changes: Dict[str, bool] = {}
        for display, source_column in ERROR_CONTEXT_COLUMNS:
            value = row[source_column]
            context_values[display] = value
            if orig_row is None:
                context_changes[display] = normalize_string(value) is not None
            else:
                context_changes[display] = not values_equal(
                    orig_row[source_column], value, source_column
                )
        errors.register_row_context(
            updated_sheet,
            row_number,
            context_values,
            context_changes,
            isolate_id if isinstance(isolate_id, str) else None,
        )


def convert_original_value(sql_column: str, vmr_value: object) -> Optional[object]:
    if sql_column in INT_COLUMNS:
        return normalize_int_like(vmr_value)
    if sql_column == "isolate_type":
        return normalize_isolate_type(vmr_value)
    return normalize_string(vmr_value)


def convert_value(
    sql_column: str,
    vmr_value: object,
    vmr_column: str,
    workbook_name: str,
    sheet_name: str,
    row_number: int,
    errors: ErrorCollector,
) -> object:
    text = normalize_string(vmr_value)
    if sql_column == "taxnode_id":
        if text is None:
            return None
        match = re.fullmatch(r"ICTV(\d+)", text.upper())
        if not match:
            errors.add(
                workbook_name,
                sheet_name,
                row_number,
                "Column 'ICTV_ID' must resemble ICTV######## to derive taxnode_id.",
            )
            return INVALID_VALUE
        return int(match.group(1))
    if sql_column in INT_COLUMNS:
        result = normalize_int_like(vmr_value)
        if result is None and text is not None:
            errors.add(
                workbook_name,
                sheet_name,
                row_number,
                f"Column '{vmr_column}' must contain an integer value.",
            )
            return INVALID_VALUE
        return result
    if sql_column == "isolate_type":
        if text is None:
            return None
        result = normalize_isolate_type(vmr_value)
        if result is None:
            errors.add(
                workbook_name,
                sheet_name,
                row_number,
                f"Column '{vmr_column}' must contain 'E' or 'A'.",
            )
            return INVALID_VALUE
        return result
    return text


def build_update_entries(
    updated_df: pd.DataFrame,
    original_df: pd.DataFrame,
    workbook_name: str,
    updated_sheet: str,
    errors: ErrorCollector,
    *,
    strict_accession: bool,
) -> List[UpdateEntry]:
    entries: List[UpdateEntry] = []
    updated_existing = updated_df[
        (updated_df["__isolate_id"].notna()) & (~updated_df["__abolished"])
    ].set_index("__isolate_id")
    original_existing = original_df[original_df["__isolate_id"].notna()].set_index("__isolate_id")

    for isolate_id, orig_row in original_existing.iterrows():
        if isolate_id not in updated_existing.index:
            continue
        upd_row = updated_existing.loc[isolate_id]
        if isinstance(upd_row, pd.DataFrame):
            continue
        changes: List[Tuple[str, Optional[object]]] = []
        original_values: List[Tuple[str, Optional[object]]] = []
        invalid = False
        for vmr_column, sql_column in UPDATABLE_TO_SQL.items():
            orig_value = orig_row[vmr_column]
            upd_value = upd_row[vmr_column]
            if values_equal(orig_value, upd_value, vmr_column):
                continue
            if vmr_column == "Virus GENBANK accession":
                change_type = classify_accession_change(orig_value, upd_value)
                if strict_accession or change_type == "meaningful":
                    errors.add(
                        workbook_name,
                        updated_sheet,
                        int(upd_row["__row_number"]),
                        (
                            "Virus GENBANK accession changed from "
                            f"'{normalize_string(orig_value) or ''}' to "
                            f"'{normalize_string(upd_value) or ''}' for isolate {isolate_id}."
                        ),
                        severity="WARNING",
                    )
            converted = convert_value(
                sql_column,
                upd_value,
                vmr_column,
                workbook_name,
                updated_sheet,
                int(upd_row["__row_number"]),
                errors,
            )
            if converted is INVALID_VALUE:
                invalid = True
                continue
            original_converted = convert_original_value(sql_column, orig_value)
            changes.append((sql_column, converted))
            original_values.append((sql_column, original_converted))
        if invalid or not changes:
            continue
        numeric_id = extract_isolate_numeric(isolate_id)
        if numeric_id is None:
            errors.add(
                workbook_name,
                updated_sheet,
                int(upd_row["__row_number"]),
                f"Isolate ID '{isolate_id}' cannot be converted to numeric form.",
            )
            continue
        entries.append(
            UpdateEntry(
                isolate_id=isolate_id,
                numeric_id=numeric_id,
                row_number=int(upd_row["__row_number"]),
                assignments=changes,
                original_values=original_values,
            )
        )
    return entries


def build_insert_entries(
    updated_df: pd.DataFrame,
    workbook_name: str,
    updated_sheet: str,
    errors: ErrorCollector,
) -> List[InsertEntry]:
    entries: List[InsertEntry] = []
    new_rows = updated_df[
        (updated_df["__isolate_id"].isna()) & (~updated_df["__abolished"])
    ]
    for _, row in new_rows.iterrows():
        row_number = int(row["__row_number"])
        values: List[Tuple[str, Optional[object]]] = []
        invalid = False
        for sql_column, vmr_column in INSERT_COLUMN_MAPPING:
            converted = convert_value(
                sql_column,
                row[vmr_column],
                vmr_column,
                workbook_name,
                updated_sheet,
                row_number,
                errors,
            )
            if converted is INVALID_VALUE:
                invalid = True
            values.append((sql_column, None if converted is INVALID_VALUE else converted))
        if invalid:
            continue
        entries.append(InsertEntry(row_number=row_number, values=values))
    return entries


def build_delete_entries(
    updated_df: pd.DataFrame,
    workbook_name: str,
    updated_sheet: str,
    errors: ErrorCollector,
) -> List[DeleteEntry]:
    entries: List[DeleteEntry] = []
    delete_rows = updated_df[updated_df["__abolished"]]
    for _, row in delete_rows.iterrows():
        row_number = int(row["__row_number"])
        isolate_id = normalize_isolate_id(row["Isolate ID"])
        if not isolate_id:
            errors.add(
                workbook_name,
                updated_sheet,
                row_number,
                "Row marked for abolish must contain an Isolate ID.",
            )
            continue
        target_value = normalize_string(row["Isolate ID"])
        if not target_value:
            errors.add(
                workbook_name,
                updated_sheet,
                row_number,
                "Row marked for abolish must contain an Isolate ID value.",
            )
            continue
        details: List[Tuple[str, Optional[object]]] = []
        for column in REQUIRED_COLUMNS:
            if column == "Species Sort":
                continue
            details.append((column, row[column]))
        numeric_id = extract_isolate_numeric(isolate_id)
        if numeric_id is None:
            errors.add(
                workbook_name,
                updated_sheet,
                row_number,
                f"Isolate ID '{isolate_id}' cannot be converted to numeric form.",
            )
            continue
        entries.append(
            DeleteEntry(
                isolate_id=isolate_id,
                numeric_id=numeric_id,
                target_value=target_value,
                row_number=row_number,
                details=details,
            )
        )
    return entries


def generate_sql_header(
    workbook_path: Path, version: str, warning: Optional[str] = None
) -> List[str]:
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S %Z")
    lines = [
        f"-- Source workbook: {workbook_path}",
        f"-- Generated: {timestamp}",
        f"-- Script version: {version}",
    ]
    if warning:
        lines.append(warning)
    lines.append("")
    return lines


def build_procedure_sql_text(
    workbook_path: Path,
    version: str,
    procedure_call: str,
    warning: Optional[str] = None,
) -> str:
    lines = generate_sql_header(workbook_path, version, warning)
    lines.append(procedure_call)
    return "\n".join(lines).rstrip() + "\n"


def format_sql_value(column: str, value: Optional[object]) -> str:
    if value is None:
        return "NULL"
    if column in INT_COLUMNS and isinstance(value, Real):
        return str(int(value))
    if isinstance(value, Integral):
        return str(int(value))
    text = str(value).replace("'", "''")
    return f"'{text}'"


def format_sql_condition(column: str, value: Optional[object]) -> str:
    if value is None:
        return f"{column} IS NULL"
    return f"{column} = {format_sql_value(column, value)}"


def build_update_sql_text(
    entries: List[UpdateEntry],
    workbook_path: Path,
    version: str,
    warning: Optional[str] = None,
) -> str:
    lines = generate_sql_header(workbook_path, version, warning)
    if not entries:
        lines.append("-- No updates required.")
        return "\n".join(lines) + "\n"
    for entry in entries:
        lines.append(f"-- {entry.isolate_id} (worksheet row {entry.row_number})")
        lines.append("UPDATE species_isolates")
        lines.append("SET")
        assignments = [
            f"    {column} = {format_sql_value(column, value)}"
            for column, value in entry.assignments
        ]
        lines.append(",\n".join(assignments))
        where_prefix = f"WHERE isolate_id = {entry.numeric_id}"
        if entry.original_values:
            conditions = [
                format_sql_condition(column, value)
                for column, value in entry.original_values
            ]
            if len(conditions) == 1:
                lines.append(f"{where_prefix} AND {conditions[0]};")
            else:
                lines.append(f"{where_prefix} AND (")
                lines.append("    " + "\n    OR ".join(conditions))
                lines.append(");")
        else:
            lines.append(f"{where_prefix};")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def build_insert_sql_text(
    entries: List[InsertEntry],
    workbook_path: Path,
    version: str,
    warning: Optional[str] = None,
) -> str:
    lines = generate_sql_header(workbook_path, version, warning)
    if not entries:
        lines.append("-- No inserts required.")
        return "\n".join(lines) + "\n"
    for entry in entries:
        lines.append(f"-- Worksheet row {entry.row_number}")
        columns = [column for column, _ in entry.values]
        values = [format_sql_value(column, value) for column, value in entry.values]
        lines.append("INSERT INTO species_isolates (")
        lines.append("    " + ",\n    ".join(columns))
        lines.append(") VALUES (")
        lines.append("    " + ",\n    ".join(values))
        lines.append(");")
        lines.append("")
    lines.extend(
        [
            "--",
            "-- update inserted species taxnode_id",
            "-- based on species_name",
            "--",
            "",
            "UPDATE species_isolates si SET",
            "        taxnode_id=(",
            "        -- get taxnode_id for latest MSL (support older species names)",
            "        select taxnode_id",
            "        from taxonomy_node_names as pt",
            "        where pt.msl_release_num = (select max(msl_release_num) from taxonomy_toc)",
            "        and pt.ictv_id = (",
            "            -- in case of older species name, get ICTV_id",
            "            select max(pit.ictv_id)",
            "            from taxonomy_node_names as pit",
            "            where pit.name=si.species_name)",
            "        )",
            "WHERE si.species_name <> 'abolished'",
            "AND si.taxnode_id is NULL",
            ";",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


def build_column_value_insert_sql_text(
    entries: List[ColumnValueInsertEntry],
    workbook_path: Path,
    version: str,
    warning: Optional[str] = None,
) -> str:
    lines = generate_sql_header(workbook_path, version, warning)
    if not entries:
        lines.append("-- No column value inserts required.")
        return "\n".join(lines) + "\n"
    for entry in entries:
        target = COLUMN_VALUE_TARGETS.get(entry.column.lower())
        if target is None:
            lines.append(
                (
                    f"-- No database mapping configured for column '{entry.column}' with "
                    f"value '{entry.value}'; skipping."
                )
            )
            lines.append("")
            continue
        row_text_values = [str(row) for row in entry.rows]
        row_text = ", ".join(row_text_values) if row_text_values else "unknown"
        lines.append(
            (
                f"-- Column '{entry.column}' value '{entry.value}' from 'Column Values' "
                f"row(s) {row_text}"
            )
        )
        lines.append(f"INSERT INTO {target.table} (")
        lines.append("    " + ",\n    ".join(target.columns))
        lines.append(") VALUES (")
        values = [format_sql_value(column, entry.value) for column in target.columns]
        lines.append("    " + ",\n    ".join(values))
        lines.append(");")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def build_delete_sql_text(
    entries: List[DeleteEntry],
    workbook_path: Path,
    version: str,
    warning: Optional[str] = None,
) -> str:
    lines = generate_sql_header(workbook_path, version, warning)
    if not entries:
        lines.append("-- No deletes required.")
        return "\n".join(lines) + "\n"
    for entry in entries:
        lines.append(f"-- {entry.isolate_id} (worksheet row {entry.row_number})")
        for column, value in entry.details:
            comment_value = normalize_string(value)
            if comment_value is None:
                comment_value = ""
            else:
                comment_value = comment_value.replace("\n", " | ")
            lines.append(f"-- {column}: {comment_value}")
        lines.append("UPDATE species_isolates SET taxnode_id=NULL, species_name='abolished' ")
        lines.append(
            f"WHERE isolate_id = {entry.numeric_id};"
        )
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def process_workbook(
    workbook_path: Path,
    workbook_name: str,
    errors: ErrorCollector,
    *,
    strict_accession: bool,
    vmr_export_path: Optional[Path] = None,
    taxonomy_paths: Optional[TaxonomyReferencePaths] = None,
) -> ProcessResult:
    with pd.ExcelFile(workbook_path) as xl:
        sheet_names = xl.sheet_names

    updated_sheet = determine_updated_sheet(sheet_names, workbook_name, errors)
    if updated_sheet is None or "Original" not in sheet_names:
        if "Original" not in sheet_names:
            errors.add(
                workbook_name,
                "",
                None,
                "Workbook does not contain an 'Original' worksheet.",
            )
        return ProcessResult(updated_sheet, [], [], [], [])

    column_constraints: Dict[str, ColumnConstraint] = {}
    column_value_inserts: List[ColumnValueInsertEntry] = []
    if "Column Values" in sheet_names:
        (
            column_constraints,
            column_value_inserts,
        ) = read_column_value_constraints(workbook_path, workbook_name, errors)
    else:
        errors.add(
            workbook_name,
            "Column Values",
            None,
            "Workbook does not contain a 'Column Values' worksheet; column value validation skipped.",
            severity="WARNING",
        )

    updated_headers = pd.read_excel(workbook_path, sheet_name=updated_sheet, nrows=0).columns
    validate_headers(workbook_name, updated_sheet, list(updated_headers), errors)

    original_headers = pd.read_excel(workbook_path, sheet_name="Original", nrows=0).columns
    validate_headers(workbook_name, "Original", list(original_headers), errors)

    updated_df = prepare_dataframe(workbook_path, updated_sheet)
    updated_df["__abolished"] = updated_df["Species Sort"].apply(is_abolish_value)
    original_df = prepare_dataframe(workbook_path, "Original")

    taxonomy_reference_sets: Dict[str, Set[str]] = {}
    if taxonomy_paths is not None:
        taxonomy_reference_sets = load_taxonomy_reference_sets(taxonomy_paths, errors)
        if taxonomy_reference_sets:
            if "Original Column Values" in sheet_names:
                original_column_values_df = pd.read_excel(
                    workbook_path, sheet_name="Original Column Values", header=0
                )
                check_original_column_values_against_taxonomy(
                    original_column_values_df, workbook_name, errors, taxonomy_reference_sets
                )
            else:
                errors.add(
                    workbook_name,
                    "Original Column Values",
                    None,
                    "Workbook does not contain an 'Original Column Values' worksheet.",
                )

    check_column_value_constraints(
        updated_df,
        workbook_name,
        updated_sheet,
        errors,
        column_constraints,
    )
    check_original_ids(original_df, workbook_name, errors)
    check_isolate_ids(updated_df, original_df, workbook_name, updated_sheet, errors)
    if vmr_export_path is not None:
        compare_original_to_vmr_export(original_df, workbook_name, errors, vmr_export_path)
    abolished_ids = {
        isolate_id
        for isolate_id in updated_df.loc[updated_df["__abolished"], "__isolate_id"].dropna()
    }
    enforce_read_only(
        updated_df,
        original_df,
        workbook_name,
        updated_sheet,
        errors,
        abolished_ids,
    )
    check_new_record_accessions(updated_df, workbook_name, updated_sheet, errors)

    delete_entries = build_delete_entries(updated_df, workbook_name, updated_sheet, errors)
    update_entries = build_update_entries(
        updated_df,
        original_df,
        workbook_name,
        updated_sheet,
        errors,
        strict_accession=strict_accession,
    )
    insert_entries = build_insert_entries(updated_df, workbook_name, updated_sheet, errors)

    register_error_context(errors, updated_df, original_df, updated_sheet)

    return ProcessResult(
        updated_sheet,
        delete_entries,
        update_entries,
        insert_entries,
        column_value_inserts,
    )


def write_sql_outputs(
    output_dir: Path,
    args: argparse.Namespace,
    result: ProcessResult,
    workbook_path: Path,
    version: str,
    had_errors: bool,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    deletes_path = output_dir / args.deletes_sql
    updates_path = output_dir / args.updates_sql
    inserts_path = output_dir / args.inserts_sql
    column_values_path = output_dir / args.column_values_sql
    update_sorts_path = output_dir / args.update_sorts_sql
    qc_sps_path = output_dir / args.qc_sps_sql
    workbook_display = workbook_path.resolve()

    warning_text: Optional[str]
    if had_errors:
        warning_text = "Errors encountered; SQL generation may be incorrect."
    else:
        warning_text = None

    deletes_sql = build_delete_sql_text(
        result.delete_entries, workbook_display, version, warning_text
    )
    updates_sql = build_update_sql_text(
        result.update_entries, workbook_display, version, warning_text
    )
    inserts_sql = build_insert_sql_text(
        result.insert_entries, workbook_display, version, warning_text
    )
    column_values_sql = build_column_value_insert_sql_text(
        result.column_value_inserts, workbook_display, version, warning_text
    )
    update_sorts_sql = build_procedure_sql_text(
        workbook_display, version, "CALL species_isolates_update_sorts();", warning_text
    )
    qc_sps_sql = build_procedure_sql_text(
        workbook_display, version, "CALL QC_run_modules(:module_filter);", warning_text
    )
    deletes_path.write_text(deletes_sql, encoding="utf-8")
    updates_path.write_text(updates_sql, encoding="utf-8")
    inserts_path.write_text(inserts_sql, encoding="utf-8")
    column_values_path.write_text(column_values_sql, encoding="utf-8")
    update_sorts_path.write_text(update_sorts_sql, encoding="utf-8")
    qc_sps_path.write_text(qc_sps_sql, encoding="utf-8")


def main() -> None:
    args = parse_args()
    workbook_path = Path(args.workbook).expanduser()
    if args.output_dir:
        output_dir = Path(args.output_dir).expanduser()
    else:
        output_dir = Path(Path(args.workbook).name).with_suffix("")

    command_line = " ".join(shlex.quote(arg) for arg in sys.argv)
    run_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    try:
        subprocess.run(["./version_git.sh"], check=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except FileNotFoundError:
        pass
    version = read_version()
    errors = ErrorCollector(
        keep_going=args.keep_going,
        command=command_line,
        version=version,
        run_date=run_timestamp,
    )
    vmr_export_path = Path(args.vmr_export).expanduser() if args.vmr_export else None
    taxonomy_paths = TaxonomyReferencePaths(
        genome_coverage=Path(args.taxonomy_genome_coverage).expanduser()
        if args.taxonomy_genome_coverage
        else None,
        molecule=Path(args.taxonomy_molecule).expanduser() if args.taxonomy_molecule else None,
        host_source=Path(args.taxonomy_host_source).expanduser()
        if args.taxonomy_host_source
        else None,
    )
    result = ProcessResult(None, [], [], [], [])

    if not workbook_path.exists():
        errors.add(
            workbook_path.name,
            "",
            None,
            f"Workbook not found: {workbook_path}",
        )
    else:
        try:
            result = process_workbook(
                workbook_path,
                workbook_path.name,
                errors,
                strict_accession=args.strict_accession,
                vmr_export_path=vmr_export_path,
                taxonomy_paths=taxonomy_paths,
            )
        except ProcessingHalted:
            pass
        except Exception as exc:  # pragma: no cover - defensive programming
            errors.extend_with_exception(workbook_path.name, exc)

    errors.write_excel(output_dir / args.errors_xlsx)
    write_sql_outputs(output_dir, args, result, workbook_path, version, errors.has_errors())

    if errors.has_errors():
        sys.exit(1)


if __name__ == "__main__":
    main()
