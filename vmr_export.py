#!/usr/bin/env python3
"""Export VMR editor/public workbooks from ICTV taxonomy data.

Design goals:
- Keep source-reading logic reusable for future tools.
- Keep workbook-writing logic data-driven from headers where practical.
- Always emit errors.xlsx with INFO/WARNING/ERROR (and SUCCESS in verbose mode).
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import re
import subprocess
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

DEFAULT_DATA_SOURCE = "test_data/export/ICTVdatabase/data"
DEFAULT_TEMPLATE = "test_data/export/template-VMR.editor.xlsx"
DEFAULT_OUTPUT = "test_out/export/VMR.editor.xlsx"
ERRORS_XLSX_NAME = "errors.xlsx"

REQUIRED_SHEETS = [
    "Version",
    r"VMR MSL.*",
    "Column definitions",
    "Column Values",
    "README.editor",
    "CHANGELOG.editor",
    "Original",
    "Original Column Values",
]

COLUMN_VALUE_SOURCES: Dict[str, Tuple[str, str] | List[str]] = {
    "Exemplar or additional isolate": ["E", "A"],
    "Genome coverage": ("taxonomy_genome_coverage", "name"),
    "Genome": ("taxonomy_molecule", "abbrev"),
    "Host source": ("taxonomy_host_source", "host_source"),
}
TABLE_SORT_KEYS: Dict[str, Tuple[str, bool]] = {
    "taxonomy_genome_coverage": ("name", True),
    "taxonomy_molecule": ("abbrev", True),
    "taxonomy_host_source": ("host_source", True),
}
VMR_SORT_TABLES = {"vmr_export", "species_isolates"}


@dataclass
class LogEntry:
    level: str
    message: str


class RunLogger:
    def __init__(self, verbose: bool):
        self.verbose = verbose
        self.entries: List[LogEntry] = []

    def _add(self, level: str, message: str, print_in_verbose: bool = True) -> None:
        self.entries.append(LogEntry(level, message))
        if level in {"ERROR", "WARNING"} or (self.verbose and print_in_verbose):
            print(f"[{level}] {message}")

    def info(self, message: str) -> None:
        self._add("INFO", message)

    def warning(self, message: str) -> None:
        self._add("WARNING", message)

    def error(self, message: str) -> None:
        self._add("ERROR", message)

    def success(self, message: str) -> None:
        if self.verbose:
            self._add("SUCCESS", message)

    def write_errors_xlsx(self, output_editor: Path) -> None:
        output_editor.parent.mkdir(parents=True, exist_ok=True)
        path = output_editor.parent / ERRORS_XLSX_NAME
        wb = Workbook()
        ws = wb.active
        ws.title = "errors"
        ws.append(["level", "message"])
        for e in self.entries:
            ws.append([e.level, e.message])
        wb.save(path)


@dataclass(frozen=True)
class MaskRule:
    table_name: str
    column: str
    value: str


def sql_like_to_regex(pattern: str) -> re.Pattern[str]:
    parts: List[str] = ["^"]
    for ch in pattern:
        if ch == "%":
            parts.append(".*")
        elif ch == "_":
            parts.append(".")
        else:
            parts.append(re.escape(ch))
    parts.append("$")
    return re.compile("".join(parts))


def load_mask_rules(mask_path: Optional[str], logger: RunLogger) -> List[MaskRule]:
    if not mask_path:
        return []
    path = Path(mask_path)
    if not path.exists():
        logger.warning(f"Mask file not found, no masking applied: {path}")
        return []
    rules: List[MaskRule] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        expected = ["table_name", "column", "value"]
        if reader.fieldnames != expected:
            raise ValueError(
                f"Mask file must have header columns: {', '.join(expected)}"
            )
        for i, row in enumerate(reader, start=2):
            table_name = str(row.get("table_name", "")).strip()
            column = str(row.get("column", "")).strip()
            value = str(row.get("value", "")).strip()
            if not table_name or not column or not value:
                raise ValueError(f"Mask file row {i} must provide table_name, column, and value")
            rules.append(MaskRule(table_name=table_name, column=column, value=value))
    logger.info(f"Loaded {len(rules)} mask rule(s) from {path}")
    return rules


class DataSourceReader:
    """Reusable source reader for flatfiles or MariaDB connection URLs."""

    def __init__(self, source: str, logger: RunLogger, mask_rules: Optional[List[MaskRule]] = None):
        self.source = source
        self.logger = logger
        self.mask_rules = mask_rules or []

    @staticmethod
    def is_db_url(value: str) -> bool:
        return value.startswith(("mysql://", "mariadb://"))

    def _read_flatfile(self, table_name: str) -> List[Dict[str, str]]:
        path = Path(self.source) / f"{table_name}.utf8.txt"
        if not path.exists():
            self.logger.warning(f"Flatfile not found for table '{table_name}': {path}")
            return []
        with path.open("r", encoding="utf-8", newline="") as handle:
            rows = [dict(row) for row in csv.DictReader(handle, delimiter="\t")]
        rows = self._apply_flatfile_masks(table_name, rows)
        self.logger.info(f"Read {len(rows)} rows from file {path}")
        return rows

    @staticmethod
    def _as_int(value: str) -> int:
        return int(str(value or "0").strip())

    def _read_db(self, table_name: str) -> List[Dict[str, str]]:
        import pandas as pd  # type: ignore
        from sqlalchemy import create_engine, text  # type: ignore

        engine = create_engine(self.source)
        rules = [rule for rule in self.mask_rules if rule.table_name == table_name]
        query = f"SELECT * FROM `{self._safe_identifier(table_name)}`"
        params: Dict[str, str] = {}
        if rules:
            clauses: List[str] = []
            for idx, rule in enumerate(rules):
                col = self._safe_identifier(rule.column)
                key = f"mask_{idx}"
                clauses.append(f"`{col}` LIKE :{key}")
                params[key] = rule.value
            query += " WHERE NOT (" + " OR ".join(clauses) + ")"
        with engine.begin() as conn:
            frame = pd.read_sql(text(query), conn, params=params)
        rows = frame.fillna("").astype(str).to_dict("records")
        if rules:
            self.logger.info(f"Applied {len(rules)} DB mask rule(s) to table/view {table_name}")
        self.logger.info(f"Read {len(rows)} rows from DB table/view {table_name}")
        return rows

    @staticmethod
    def _safe_identifier(value: str) -> str:
        text = str(value).strip()
        if not re.fullmatch(r"[A-Za-z0-9_]+", text):
            raise ValueError(f"Unsafe SQL identifier in mask/table name: {value!r}")
        return text

    def _apply_flatfile_masks(self, table_name: str, rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
        rules = [rule for rule in self.mask_rules if rule.table_name == table_name]
        if not rules:
            return rows
        compiled = [
            (rule.column, sql_like_to_regex(rule.value))
            for rule in rules
        ]
        filtered: List[Dict[str, str]] = []
        for row in rows:
            masked = False
            for column, pattern in compiled:
                value = str(row.get(column, "") or "")
                if pattern.match(value):
                    masked = True
                    break
            if not masked:
                filtered.append(row)
        self.logger.info(f"Applied {len(rules)} flatfile mask rule(s) to table '{table_name}'")
        return filtered

    def read_table(self, table_name: str) -> List[Dict[str, str]]:
        if os.path.isdir(self.source):
            rows = self._read_flatfile(table_name)
        elif self.is_db_url(self.source):
            rows = self._read_db(table_name)
        else:
            raise ValueError("Unsupported data source")

        sort_spec = TABLE_SORT_KEYS.get(table_name)
        if sort_spec:
            key_name, case_insensitive = sort_spec
            if case_insensitive:
                rows.sort(key=lambda row: str(row.get(key_name, "") or "").strip().lower())
            else:
                rows.sort(key=lambda row: str(row.get(key_name, "") or "").strip())
        elif table_name in VMR_SORT_TABLES:
            rows.sort(
                key=lambda row: (
                    self._as_int(row.get("Species Sort", "0")),
                    self._as_int(row.get("Isolate Sort", "0")),
                )
            )
        return rows


def parse_args(argv: Optional[Iterable[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export VMR data from DB/flatfile into template workbook.")
    parser.add_argument(
        "-i", "--data_source", default=DEFAULT_DATA_SOURCE,
        help='Can be either a MariaDB URL (e.g. "mysql://root:secret@localhost:3306/ictv_taxonomy") or a flatfile directory.',
    )
    parser.add_argument(
        "-t", "--template", default=DEFAULT_TEMPLATE,
        help="Template workbook used for formatting. Data content is replaced.",
    )
    parser.add_argument(
        "-o", "--output", default=DEFAULT_OUTPUT,
        help="Must end with .editor.xlsx; writes FILEPATH.editor.xlsx and FILEPATH.xlsx",
    )
    parser.add_argument(
        "--mask",
        default="db_mask.tsv",
        help="TSV with columns table_name, column, value (SQL LIKE pattern) for source masking.",
    )
    parser.add_argument("-v", "--verbose", action="store_true", help="Verbose mode prints INFO/SUCCESS logs.")
    parser.add_argument(
        "-k",
        "--keep-going",
        action="store_true",
        help="Treat Column definitions validation mismatches as warnings and continue.",
    )
    return parser.parse_args(list(argv) if argv is not None else None)


def find_vmr_sheet_name(sheetnames: Sequence[str]) -> Optional[str]:
    for name in sheetnames:
        if re.fullmatch(r"VMR MSL.*", name):
            return name
    return None


def validate_template(wb, logger: RunLogger) -> Optional[str]:
    vmr_sheet = find_vmr_sheet_name(wb.sheetnames)
    missing: List[str] = []
    for required in REQUIRED_SHEETS:
        if required == r"VMR MSL.*":
            if vmr_sheet is None:
                missing.append(required)
            else:
                logger.success(f"Validated worksheet: {vmr_sheet}")
        elif required not in wb.sheetnames:
            missing.append(required)
        else:
            logger.success(f"Validated worksheet: {required}")
    if missing:
        logger.error(f"Template is missing worksheet(s): {', '.join(missing)}")
        return None
    return vmr_sheet


def header_map(ws) -> Dict[str, int]:
    return {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value not in (None, "")}


def clear_sheet_data(ws, start_row: int = 2, start_col: int = 1, end_col: Optional[int] = None) -> None:
    if end_col is None:
        end_col = ws.max_column
    for r in range(start_row, ws.max_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(r, c).value = None


def apply_column_values(ws, reader: DataSourceReader, logger: RunLogger) -> None:
    clear_sheet_data(ws, start_row=2)
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for col_idx, hdr in enumerate(headers, start=1):
        key = str(hdr).strip() if hdr else ""
        source = COLUMN_VALUE_SOURCES.get(key)
        if source is None:
            continue
        if isinstance(source, list):
            values = source
        else:
            table_name, col_name = source
            values = [str(r.get(col_name, "")).strip() for r in reader.read_table(table_name)]
            values = [v for v in values if v]
        for row_idx, value in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx).value = value
        logger.info(f"Wrote {len(values)} values to sheet '{ws.title}' column '{key}'")


def autofit_sheet_columns(ws, max_width: int = 80) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            text = str(value)
            cell_len = max((len(line) for line in text.splitlines()), default=0)
            if cell_len > max_len:
                max_len = cell_len
        width = min(max_width, max(8, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _parse_heading_tokens(heading: str) -> List[str]:
    return [part.strip() for part in heading.split(",") if part and part.strip()]


def _extract_definition_bullets(text: str) -> List[str]:
    bullets: List[str] = []
    for line in text.splitlines():
        match = re.match(r"^\s*-\s*(.+?)\s*$", line)
        if match:
            bullets.append(match.group(1).strip())
    return bullets


def _normalize_match_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value)
    out: List[str] = []
    for ch in normalized:
        cat = unicodedata.category(ch)
        if ch.isspace() or cat.startswith("Z") or cat.startswith("P"):
            continue
        out.append(ch.lower())
    return "".join(out)


def _choose_match(source: str, candidates: List[str]) -> Tuple[str, bool]:
    if source in candidates:
        return source, True
    return candidates[0], False


def validate_column_definitions(vmr_ws, defs_ws, column_values_ws) -> Tuple[List[str], List[str]]:
    issues: List[str] = []
    mappings: List[str] = []
    vmr_headers = header_map(vmr_ws)
    editor_notes_col = vmr_headers.get("Editor Notes")
    if not editor_notes_col:
        issues.append(
            f"Worksheet '{vmr_ws.title}' is missing 'Editor Notes' header needed for definitions validation"
        )
        return issues, mappings

    defs_headers = header_map(defs_ws)
    defs_col_col = defs_headers.get("Column")
    defs_heading_col = defs_headers.get("Heading")
    defs_definition_col = defs_headers.get("Definition")
    if not all([defs_col_col, defs_heading_col, defs_definition_col]):
        missing: List[str] = []
        if not defs_col_col:
            missing.append("Column")
        if not defs_heading_col:
            missing.append("Heading")
        if not defs_definition_col:
            missing.append("Definition")
        issues.append(
            f"Worksheet '{defs_ws.title}' missing required column(s) for validation: {', '.join(missing)}"
        )
        return issues, mappings

    heading_raw_by_norm: Dict[str, List[str]] = {}
    heading_defs_by_norm: Dict[str, List[str]] = {}
    range_dr_raw_tokens: List[str] = []
    range_dr_by_norm: Dict[str, List[str]] = {}
    found_dr_row = False
    for r in range(2, defs_ws.max_row + 1):
        column_text = str(defs_ws.cell(r, defs_col_col).value or "").strip()
        heading_text = str(defs_ws.cell(r, defs_heading_col).value or "").strip()
        definition_text = str(defs_ws.cell(r, defs_definition_col).value or "").strip()
        if not column_text and not heading_text and not definition_text:
            continue
        tokens = _parse_heading_tokens(heading_text)
        for token in tokens:
            norm_token = _normalize_match_text(token)
            if not norm_token:
                continue
            heading_raw_by_norm.setdefault(norm_token, [])
            if token not in heading_raw_by_norm[norm_token]:
                heading_raw_by_norm[norm_token].append(token)
            heading_defs_by_norm.setdefault(norm_token, [])
            if definition_text:
                heading_defs_by_norm[norm_token].append(definition_text)
        if re.fullmatch(r"D\s*-\s*R", column_text):
            found_dr_row = True
            range_dr_raw_tokens = tokens
            range_dr_by_norm = {}
            for token in tokens:
                norm_token = _normalize_match_text(token)
                if not norm_token:
                    continue
                range_dr_by_norm.setdefault(norm_token, [])
                if token not in range_dr_by_norm[norm_token]:
                    range_dr_by_norm[norm_token].append(token)

    # All VMR columns before "Editor Notes" must be documented in Column definitions.
    vmr_required_headers: List[str] = []
    for c in range(1, editor_notes_col):
        value = vmr_ws.cell(1, c).value
        header = str(value).strip() if value else ""
        if header:
            vmr_required_headers.append(header)
    for header in vmr_required_headers:
        norm_header = _normalize_match_text(header)
        candidates = heading_raw_by_norm.get(norm_header, [])
        if not candidates:
            issues.append(
                f"Worksheet '{defs_ws.title}' is missing documentation for VMR header '{header}' from worksheet '{vmr_ws.title}'"
            )
            continue
        matched, exact = _choose_match(header, candidates)
        note = "" if exact else " (normalized match; text differs)"
        mappings.append(f"Header map: VMR '{header}' -> Defs '{matched}'{note}")

    # "Realm" through "Species" are documented via a single D-R row with CSV headings.
    vmr_dr_headers: List[str] = []
    for c in range(4, 19):  # D..R
        value = vmr_ws.cell(1, c).value
        header = str(value).strip() if value else ""
        if header:
            vmr_dr_headers.append(header)
    if not found_dr_row:
        issues.append(f"Worksheet '{defs_ws.title}' is missing the grouped 'D - R' row for Realm..Species headings")
    else:
        for header in vmr_dr_headers:
            norm_header = _normalize_match_text(header)
            candidates = range_dr_by_norm.get(norm_header, [])
            if not candidates:
                issues.append(
                    f"Worksheet '{defs_ws.title}' D - R heading list is missing '{header}' from worksheet '{vmr_ws.title}'"
                )
                continue
            matched, exact = _choose_match(header, candidates)
            note = "" if exact else " (normalized match; text differs)"
            mappings.append(f"D-R map: VMR '{header}' -> Defs '{matched}'{note}")

    # Validate that each controlled value in Column Values appears in the matching Definition bullets.
    col_values_headers = header_map(column_values_ws)
    for header in COLUMN_VALUE_SOURCES:
        vmr_col = vmr_headers.get(header)
        cv_col = col_values_headers.get(header)
        if not vmr_col or not cv_col:
            continue
        values: List[str] = []
        for r in range(2, column_values_ws.max_row + 1):
            value = str(column_values_ws.cell(r, cv_col).value or "").strip()
            if value:
                values.append(value)
        norm_header = _normalize_match_text(header)
        defs_texts = heading_defs_by_norm.get(norm_header, [])
        if not defs_texts:
            issues.append(
                f"Worksheet '{defs_ws.title}' has no Definition text for controlled-value header '{header}'"
            )
            continue
        defs_combined = "\n".join(defs_texts)
        bullets = _extract_definition_bullets(defs_combined)
        bullets_by_norm: Dict[str, List[str]] = {}
        for bullet in bullets:
            norm_bullet = _normalize_match_text(bullet)
            if not norm_bullet:
                continue
            bullets_by_norm.setdefault(norm_bullet, [])
            if bullet not in bullets_by_norm[norm_bullet]:
                bullets_by_norm[norm_bullet].append(bullet)
        for value in values:
            norm_value = _normalize_match_text(value)
            if not norm_value:
                continue
            candidates = bullets_by_norm.get(norm_value, [])
            if bullets and not candidates:
                issues.append(
                    f"Worksheet '{defs_ws.title}' Definition for '{header}' is missing value '{value}' from worksheet '{column_values_ws.title}'"
                )
                continue
            if bullets and candidates:
                matched, exact = _choose_match(value, candidates)
                note = "" if exact else " (normalized match; text differs)"
                mappings.append(f"Value map [{header}]: Column Values '{value}' -> Definition '{matched}'{note}")
            elif not bullets:
                defs_norm = _normalize_match_text(defs_combined)
                if norm_value not in defs_norm:
                    issues.append(
                        f"Worksheet '{defs_ws.title}' Definition for '{header}' is missing value '{value}' from worksheet '{column_values_ws.title}'"
                    )
                else:
                    mappings.append(
                        f"Value map [{header}]: Column Values '{value}' -> Definition text match (normalized substring)"
                    )
    return issues, mappings


def get_git_commit() -> str:
    try:
        return subprocess.check_output(["git", "rev-parse", "--short", "HEAD"], text=True).strip()
    except Exception:
        return "unknown"


def validate_vmr_columns(data_rows: List[Dict[str, str]], ws, logger: RunLogger) -> Tuple[List[str], List[str], List[str], List[str]]:
    data_headers = list(data_rows[0].keys()) if data_rows else []
    ws_headers = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else "" for c in range(1, ws.max_column + 1)]

    delta_columns = [h for h in ws_headers if "𝚫" in h]
    data_columns = []
    change_columns = []
    first_delta_index = next((i for i, h in enumerate(ws_headers) if "𝚫" in h), None)
    if first_delta_index is not None:
        data_columns = [h for h in ws_headers[:first_delta_index] if h]
        change_columns = [h for h in ws_headers[first_delta_index + len(delta_columns):] if h]

    ws_non_formula_data = [h for h in ws_headers if h and "𝚫" not in h and not h.startswith("=")]

    for col in data_headers:
        if col not in ws_non_formula_data and col != "isolate_id":
            logger.error(f"Data column '{col}' is missing from worksheet '{ws.title}'")
    for col in ws_non_formula_data:
        if col not in data_headers:
            logger.error(f"Worksheet column '{col}' is missing from data source")
        else:
            logger.success(f"Column match: '{col}'")

    if change_columns and len(change_columns) != max(0, len(data_columns) - 1):
        logger.error("Change-column count mismatch relative to data columns (excluding Isolate ID)")

    return data_headers, ws_headers, data_columns, change_columns


def write_vmr_rows(ws, data_rows: List[Dict[str, str]], data_columns: List[str], logger: RunLogger) -> int:
    clear_sheet_data(ws, start_row=2)
    for row_idx, row in enumerate(data_rows, start=2):
        for col_idx, col_name in enumerate(data_columns, start=1):
            if col_name in {"Species Sort", "Isolate Sort"}:
                value = int(str(row.get(col_name, "0") or "0").strip())
                ws.cell(row=row_idx, column=col_idx).value = value
                ws.cell(row=row_idx, column=col_idx).number_format = "0"
            else:
                ws.cell(row=row_idx, column=col_idx).value = row.get(col_name, "")
                ws.cell(row=row_idx, column=col_idx).number_format = "@"
    logger.info(f"Wrote {len(data_rows)} rows into worksheet '{ws.title}'")
    return len(data_rows)


def apply_isolate_id_hyperlinks(ws, logger: RunLogger) -> None:
    headers = header_map(ws)
    isolate_col = headers.get("Isolate ID")
    if not isolate_col:
        return
    updated = 0
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, isolate_col)
        value = str(cell.value or "").strip()
        if not value:
            continue
        if value.startswith("=HYPERLINK("):
            cell.style = "Hyperlink"
            continue
        cell.value = f'=HYPERLINK("https://ictv.global/id/{value}","{value}")'
        cell.style = "Hyperlink"
        updated += 1
    if updated:
        logger.info(f"Added {updated} isolate hyperlinks in worksheet '{ws.title}'")


def fill_original_formulas(original_ws, vmr_sheet_name: str, row_count: int, data_columns: List[str], change_columns: List[str]) -> None:
    headers = header_map(original_ws)
    count_col = headers.get("#𝚫")
    changes_col = headers.get("𝚫s")
    if not count_col or not changes_col:
        return
    change_start_col = changes_col + 1
    change_end_col = change_start_col + len(change_columns) - 1
    start_letter = get_column_letter(change_start_col)
    end_letter = get_column_letter(change_end_col)

    for r in range(2, row_count + 2):
        original_ws.cell(r, count_col).value = f'=COUNTIF({start_letter}{r}:{end_letter}{r},"?*")'
        original_ws.cell(r, changes_col).value = f'=_xlfn.TEXTJOIN(",",TRUE,{start_letter}{r}:{end_letter}{r})'

    for offset, change_name in enumerate(change_columns, start=0):
        source_idx = offset + 2  # Skip Isolate ID
        target_col = change_start_col + offset
        source_letter = get_column_letter(source_idx)
        target_letter = get_column_letter(target_col)
        if change_name.startswith("QC_"):
            continue
        for r in range(2, row_count + 2):
            if offset == 0:
                formula = (
                    f"=IF(NOT(EXACT({source_letter}{r},INDEX('{vmr_sheet_name}'!{source_letter}:{source_letter},"
                    f"MATCH($A{r},'{vmr_sheet_name}'!$A:$A,0)))),{target_letter}$1,\"\")"
                )
            else:
                formula = (
                    f"=IF(NOT(EXACT({source_letter}{r},INDEX('{vmr_sheet_name}'!{source_letter}:{source_letter},"
                    f"MATCH($A{r},'{vmr_sheet_name}'!$A:$A,0)))),_xlfn.CONCAT({target_letter}$1,\":\",{source_letter}{r}),\"\")"
                )
            original_ws.cell(r, target_col).value = formula


def fill_vmr_delta_formulas(vmr_ws, row_count: int, logger: RunLogger) -> None:
    headers = header_map(vmr_ws)
    changes_col = headers.get("𝚫s")
    if not changes_col:
        logger.error(f"Worksheet '{vmr_ws.title}' is missing column header '𝚫s' needed for delta formulas")
        return
    count_col = changes_col - 1
    if count_col < 1:
        logger.error(f"Worksheet '{vmr_ws.title}' does not have a column before '𝚫s' for '#𝚫' formulas")
        return
    count_header = vmr_ws.cell(1, count_col).value
    if count_header is None or "#𝚫" not in str(count_header):
        logger.error(f"Worksheet '{vmr_ws.title}' expected '#𝚫' header before '𝚫s', found '{count_header}'")
        return
    count_letter = get_column_letter(count_col)
    changes_letter = get_column_letter(changes_col)
    count_alignment = Alignment(horizontal="right")
    for r in range(2, row_count + 2):
        count_cell = vmr_ws.cell(r, count_col)
        count_cell.value = f'=IF(AND(LEN($A{r})=0,LEN($R{r})>0),"1",INDEX(Original!{count_letter}:{count_letter},MATCH($A{r},Original!$A:$A,0)))'
        count_cell.number_format = "0"
        count_cell.alignment = count_alignment
        vmr_ws.cell(r, changes_col).value = f'=IF(AND(LEN($A{r})=0,LEN($R{r})>0),"new isolate",INDEX(Original!{changes_letter}:{changes_letter},MATCH($A{r},Original!$A:$A,0)))'


def update_changelog(wb, msl_release_num: str, version_tag: str, cli_args: Sequence[str]) -> None:
    ws = wb["CHANGELOG.editor"]
    headers = header_map(ws)
    when_col = headers.get("When", 1)
    who_col = headers.get("Who", 2)
    what_col = headers.get("What", 3)

    next_row = ws.max_row + 1
    while next_row > 1 and all(ws.cell(next_row - 1, c).value in (None, "") for c in [when_col, who_col, what_col]):
        next_row -= 1
    if any(ws.cell(next_row, c).value not in (None, "") for c in [when_col, who_col, what_col]):
        next_row += 1

    today = dt.date.today().isoformat()
    ws.cell(next_row, when_col).value = today
    ws.cell(next_row, who_col).value = f"vmr_export.py@{get_git_commit()}"
    repo_root = Path(__file__).resolve().parent
    def _git_or_unknown(args: Sequence[str]) -> str:
        try:
            return subprocess.check_output(args, text=True, cwd=repo_root).strip()
        except Exception:
            return "unknown"
    origin_url = _git_or_unknown(["git", "remote", "get-url", "origin"])
    head_hash = _git_or_unknown(["git", "rev-parse", "HEAD"])
    ws.cell(next_row, what_col).value = (
        f"Command: {' '.join(cli_args)}"
        + f"\nmsl_release: MSL{msl_release_num} version {version_tag}"
        + f"\norigin: {origin_url}"
        + f"\ncommit: {head_hash}"
    )


def trim_public_workbook(editor_path: Path, public_path: Path, logger: RunLogger) -> None:
    wb = load_workbook(editor_path)
    for sheet in ["README.editor", "CHANGELOG.editor", "Original", "Original Column Values"]:
        if sheet in wb.sheetnames:
            wb.remove(wb[sheet])

    vmr = wb[find_vmr_sheet_name(wb.sheetnames)]
    headers = header_map(vmr)
    editor_notes_col = headers.get("Editor Notes")
    if editor_notes_col:
        vmr.delete_cols(editor_notes_col, vmr.max_column - editor_notes_col + 1)

    for ws in wb.worksheets:
        ws.data_validations.dataValidation = []
        ws.conditional_formatting._cf_rules = {}

    if "Version" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Version")

    wb.save(public_path)
    logger.info(f"Wrote {len(wb.sheetnames)} worksheet(s) to public file {public_path}")


def run(argv: Optional[Iterable[str]] = None) -> int:
    args = parse_args(argv)
    logger = RunLogger(args.verbose)

    output_editor = Path(args.output)
    output_public = Path(str(output_editor).replace(".editor.xlsx", ".xlsx"))

    def fail(message: str) -> int:
        logger.error(message)
        logger.write_errors_xlsx(output_editor)
        return 1

    if not str(output_editor).endswith(".editor.xlsx"):
        return fail("output must be formatted as FILEPATH.editor.xlsx")
    if not (os.path.isdir(args.data_source) or DataSourceReader.is_db_url(args.data_source)):
        return fail("data_source must be an existing directory or a MariaDB connection string")
    if not Path(args.template).exists():
        return fail(f"Template file not found: {args.template}")

    mask_rules = load_mask_rules(args.mask, logger)
    reader = DataSourceReader(args.data_source, logger, mask_rules=mask_rules)
    wb = load_workbook(args.template)
    logger.info(f"Read template workbook from {args.template}")

    # Main sheet is named "VMR MSL##" 
    # where ## is the latest msl_release_num from the taxonomy_toc table, 
    # so go find that sheet and report on it's actual name.
    vmr_sheet_name = validate_template(wb, logger)
    if vmr_sheet_name is None:
        logger.write_errors_xlsx(output_editor)
        return 1

    # get main taxonony data worksheets
    vmr_ws = wb[vmr_sheet_name]
    original_ws = wb["Original"]
    defs_ws = wb["Column definitions"]
    col_values_ws = wb["Column Values"]

    # fill in controlled vocabulary workseets
    apply_column_values(wb["Original Column Values"], reader, logger)
    apply_column_values(col_values_ws, reader, logger)
    autofit_sheet_columns(wb["Original Column Values"])
    autofit_sheet_columns(col_values_ws)
    defs_issues, defs_mappings = validate_column_definitions(vmr_ws, defs_ws, col_values_ws)
    for mapping in defs_mappings:
        logger.info(mapping)
    if defs_issues:
        if args.keep_going:
            for issue in defs_issues:
                logger.warning(issue)
        else:
            for issue in defs_issues:
                logger.error(issue)
            logger.write_errors_xlsx(output_editor)
            return 1

    # pull latest taxonomy/isolate data from files/db
    vmr_rows = reader.read_table("vmr_export")
    if not vmr_rows:
        return fail("No rows found in vmr_export source")

    data_headers, _ws_headers, data_columns, change_columns = validate_vmr_columns(vmr_rows, original_ws, logger)
    _ = data_headers
    if any(e.level == "ERROR" for e in logger.entries):
        logger.write_errors_xlsx(output_editor)
        return 1

    # put new taxonomy/isoaltes into workbook
    if args.verbose:
        logger.info("Populating Original worksheet")
    row_count = write_vmr_rows(original_ws, vmr_rows, data_columns, logger)
    fill_original_formulas(original_ws, vmr_sheet_name, row_count, data_columns, change_columns)

    if args.verbose:
        logger.info(f"Populating {vmr_sheet_name} worksheet")
    write_vmr_rows(vmr_ws, vmr_rows, data_columns, logger)
    apply_isolate_id_hyperlinks(vmr_ws, logger)
    fill_vmr_delta_formulas(vmr_ws, row_count, logger)

    # Step 4 (minimal): apply requested highlights to Column Values and VMR sheet.
    #
    # set up visuals (OpenPyXL doesn't support the "Green Fill with Dark Green Text", etc, presets)
    green_fill = PatternFill(fill_type="solid", fgColor="FFC6EFCE", bgColor="FFC6EFCE")
    green_font = Font(color="FF006100")
    red_fill = PatternFill(fill_type="solid", fgColor="FFFFC7CE", bgColor="FFFFC7CE")
    red_font = Font(color="FF9C0006")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFFEB9C", bgColor="FFFFEB9C")
    yellow_font = Font(color="FF9C6500")

    # 
    # highlight changes in WS "Column Values" relative to "Original Column Values" sheet
    #
    col_values_ws = wb["Column Values"]
    # remove all previous conditional formatting
    col_values_ws.conditional_formatting._cf_rules = {}
    col_values_ws.conditional_formatting.add(
        f"A1:{get_column_letter(col_values_ws.max_column)}1048576",
        FormulaRule(formula=["=AND(NOT(ISBLANK(A1)), ISNA(MATCH(A1, 'Original Column Values'!A:A, 0)))"], fill=red_fill, font=red_font),
    )

    # 
    # highlight changes in WS "VMR MSL##" main worksheet relative to "Original" sheet, and count/describe changes in "#𝚫"/"𝚫s" columns
    #
    # remove all previous conditional formatting
    vmr_ws.conditional_formatting._cf_rules = {}
    # find all the columns we need by header name 
    vmr_headers = header_map(vmr_ws)
    # find count_col, the header that contains the substring "#𝚫" for the count of changes
    count_col = None
    for header, col in vmr_headers.items():
        if "#𝚫" in header:
            count_col = col
            break
    changes_col = vmr_headers.get("𝚫s")
    isolate_col = vmr_headers.get("Isolate ID")
    species_col = vmr_headers.get("Species")
    if not all([count_col, changes_col, isolate_col, species_col]):
        missing = []
        if not count_col:
            missing.append("#𝚫")
        if not changes_col:
            missing.append("𝚫s")
        if not isolate_col:
            missing.append("Isolate ID")
        if not species_col:
            missing.append("Species")
        logger.error(
            f"Worksheet '{vmr_ws.title}' missing required column(s): {', '.join(missing)}"
        )
        logger.write_errors_xlsx(output_editor)
        return 1
 
    # add yellow highlights to deltas columns when there are changes in the row
    count_col_letter = get_column_letter(count_col)
    changes_col_letter = get_column_letter(changes_col)
    vmr_ws.conditional_formatting.add(f"{count_col_letter}1:{count_col_letter}1048576", FormulaRule(formula=[f"={count_col_letter}1>0"], fill=yellow_fill, font=yellow_font))
    vmr_ws.conditional_formatting.add(f"{changes_col_letter}1:{changes_col_letter}1048576", FormulaRule(formula=[f"=LEN({changes_col_letter}1)>0"], fill=yellow_fill, font=yellow_font))

    # add green highlights for new isolates (insert new row) in the VMR sheet 
    # (where "Isolate ID" is empty and "Species" is not)
    editor_notes_col= vmr_headers.get("Editor Notes")
    
    isolate_col_letter = get_column_letter(isolate_col)
    species_col_letter = get_column_letter(species_col)
    vmr_ws.conditional_formatting.add(
        f"A1:AC1048576",
        FormulaRule(formula=[f"=AND(LEN(${isolate_col_letter}1)=0, LEN(${species_col_letter}1)>0)"], fill=green_fill, font=green_font),
    )
        
    # add green highlighting for individual cell changes in the VMR sheet 
    # (where the cell value is not blank and does not match the Original sheet value for that Isolate ID)
    editor_notes_col_letter = get_column_letter(editor_notes_col)
    vmr_ws.conditional_formatting.add(
        f"A1:{editor_notes_col_letter}1048576",
        FormulaRule(formula=[f"=NOT(EXACT(A1,INDEX(Original!A:A,MATCH($A1,Original!$A:$A,0))))"], fill=green_fill, font=green_font),
    )
    
    # Freeze scroling for Row 1 and Columns A-C in the VMR sheet.
    vmr_ws.freeze_panes = "D2"
    
    # add data-validation dropdowns to certain columns in the VMR sheet, sourcing from the "Column Values" sheet
    for col_letter, list_col in [("T", "A"), ("Y", "B"), ("Z", "C"), ("AA", "D")]:
        dv = DataValidation(type="list", formula1=f"='Column Values'!${list_col}:${list_col}")
        vmr_ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}1048576")

    toc_rows = reader.read_table("taxonomy_toc")
    msl_release = ""
    version_tag = ""
    if toc_rows:
        latest = max(toc_rows, key=lambda r: int(str(r.get("msl_release_num", "0") or "0")))
        msl_release = str(latest.get("msl_release_num", "")).strip()
        version_tag = str(latest.get("version_tag", "")).strip()
        if msl_release and not vmr_sheet_name.endswith(msl_release):
            return fail(f"Template VMR sheet '{vmr_sheet_name}' does not match latest msl_release_num '{msl_release}'")

    # add this release info to the change-log sheet, along with the count_col_letterI args used to generate it (for traceability)
    update_changelog(wb, msl_release, version_tag, [sys.argv[0], *sys.argv[1:]])

    # save full (editor) version 
    if "Version" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Version")
    output_editor.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_editor)
    logger.info(f"Wrote workbook with {len(wb.sheetnames)} worksheet(s) to {output_editor}")

    # trim out non-public sheets and columns, and save public version
    trim_public_workbook(output_editor, output_public, logger)
    logger.write_errors_xlsx(output_editor)
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
